# Convert CCDI Excel files to CSV for input into BDIViz
import argparse
import openpyxl as xl
import pandas as pd
from crdclib import crdclib


def main(args):
    configs = crdclib.readYAML(args.configfile)
    
    for xlfile in configs['excelfiles']:
        temp = xlfile.split('\\')
        filename = temp[-1]
        #print(filename)
        
        #wb_df = pd.DataFrame()
        wb = xl.load_workbook(xlfile)
        sheetnames = wb.sheetnames
        #Dump the README
        sheetnames.pop(0)
        #wb_df = pd.ExcelFile(xlfile)
        #x=1
        dflist = []
        for sheet in sheetnames:
            #wb.active = wb[sheet]
            #temp_df = pd.read_excel(xlfile, sheet)
            #wb_df = wb_df.con(temp_df)
            dflist.append(pd.read_excel(xlfile, sheet))
        
        wb_df = pd.concat(dflist)
        wb_df.dropna(how='all', axis='columns',inplace=True)
        filename = filename+".csv"
        filename = configs['outputpath']+filename
        
        wb_df.to_csv(filename, sep='\t', index=False)            


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("-c", "--configfile", required=True,  help="Configuration file containing all the input info")
    parser.add_argument("-v", "--verbose", help="Verbose Output")

    args = parser.parse_args()

    main(args)